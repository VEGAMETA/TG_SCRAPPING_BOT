import openpyxl

from time import sleep
from . import AccessManager

class XlsxManager(AccessManager):
    def __init__(self, FILE_NAME: str, COLUMNS: list[str]) -> None:
        self.filename = FILE_NAME
        self.columns = COLUMNS

        self.book = openpyxl.Workbook(iso_dates=True)
        self.sheet = self.book.active
        self.row = 1
        self.error_save = False
        self.add_data(*self.columns)

    def add_data(self, *args) -> None:
        for key, arg in enumerate(args):
            self.sheet.cell(self.row, key + 1, arg)
        self.row += 1
    
    def get_data(self, row: int, col: int) -> str:
        return self.sheet.cell(row, col).value

    def __del__(self) -> None:
        while True:
            try:
                self.book.save(self.filename)
                self.book.close()
                break
            except PermissionError:
                if not self.error_save:
                    self.error_save = True
                    print(f"PermissionError: {self.filename}\nMake sure the file is not open\nTrying to save again in 5 seconds...")
                sleep(5)
